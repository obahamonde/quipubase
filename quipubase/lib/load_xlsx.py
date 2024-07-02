import json
from dataclasses import dataclass
from openpyxl import load_workbook
from ._base import RawDoc
from datetime import datetime, date, time, timedelta
from decimal import Decimal


class JsonEncoder(json.JSONEncoder):
    def default(self, o: object) -> object:
        if isinstance(o, datetime):
            return o.isoformat()
        if isinstance(o, date):
            return o.isoformat()
        if isinstance(o, time):
            return o.isoformat()
        if isinstance(o, timedelta):
            return o.total_seconds()
        if isinstance(o, Decimal):
            return str(o)
        return super().default(o)


@dataclass
class ExcelLoader(RawDoc):
    def extract_text(self):
        wb = load_workbook(filename=self.file_path, data_only=True)
        for sheet_name in self.wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        data_dict = {
                            "row": cell.row,
                            "column": cell.column,
                            "value": cell.value,
                            "sheet": sheet_name,
                        }
                        yield json.dumps(data_dict, cls=JsonEncoder)

    def extract_image(self):
        wb = load_workbook(filename=self.file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for image in sheet._images:  # type: ignore
                yield image._data()  # type: ignore
